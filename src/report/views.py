from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from django.db.models import Sum
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from student.models.student_model import StudentModel
from teacher.models.teacher_model import TeacherModel
from user.models.user_model import UserModel


# Create your views here.

@login_required
def index(request):

    if request.method == 'GET':
        select = request.GET.get('select', '')
        format_type = request.GET.get('format', '')
        if select == 'students':
            if format_type == 'excel':
                return redirect('report:students_list_to_excel')
            elif format_type == 'pdf':
                return redirect('report:students_list_to_pdf')
        elif select == 'teachers':
            if format_type == 'excel':
                return redirect('report:teachers_list_to_excel')
            elif format_type == 'pdf':
                return redirect('report:teachers_list_to_pdf')
        elif select == 'users':
            if format_type == 'excel':
                return redirect('report:users_list_to_excel')
            elif format_type == 'pdf':
                return redirect('report:users_list_to_pdf')

    return render(request,  'report/index.html')


@login_required
def students_list_to_excel(request):
    students = StudentModel.objects.all().order_by('id')

    workbook = Workbook()
    sheet = workbook.active

    sheet['A1'] = 'ID'
    sheet['B1'] = 'Nom d\'utilisateur'
    sheet['C1'] = 'Nom'
    sheet['D1'] = 'Prénoms'
    sheet['E1'] = 'Genre'
    sheet['F1'] = 'Date de naissance'
    sheet['G1'] = 'Adresse'
    sheet['H1'] = 'Numéro de téléphone'
    sheet['I1'] = 'Matricule'
    sheet['J1'] = 'Numéro de téléphone du père'

    row = 2
    for student in students:
        sheet.cell(row=row, column=1).value = student.pk
        sheet.cell(row=row, column=2).value = student.user.username
        sheet.cell(row=row, column=3).value = student.first_name
        sheet.cell(row=row, column=4).value = student.last_name
        sheet.cell(row=row, column=5).value = student.gender
        date_formatted = student.birthday.strftime("%d-%m-%Y")
        sheet.cell(row=row, column=6).value = date_formatted
        sheet.cell(row=row, column=7).value = f"{student.address.country}-{student.address.city}-{student.address.street}"
        sheet.cell(row=row, column=8).value = student.phone_number
        sheet.cell(row=row, column=9).value = student.matricule 
        sheet.cell(row=row, column=10).value = student.phone_number_father 
        
        row += 1
    total_students = students.aggregate(Sum('id'))['id__sum']
    last_row = len(students) + 2
    # sheet[f'B{last_row}'] = f"Total = {total_students}"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename='Student_list_{datetime.today().strftime('%d_%m_%Y')}.xlsx'"

    workbook.save(response)
  
    return response


@login_required
def teachers_list_to_excel(request):
    teachers = TeacherModel.objects.all().order_by('id')

    workbook = Workbook()
    sheet = workbook.active

    sheet['A1'] = 'ID'
    sheet['B1'] = 'Nom d\'utilisateur'
    sheet['C1'] = 'Nom'
    sheet['D1'] = 'Prénoms'
    sheet['E1'] = 'Genre'
    sheet['F1'] = 'Date de naissance'
    sheet['G1'] = 'Adresse'
    sheet['H1'] = 'Numéro de téléphone'
    sheet['I1'] = 'Spécialité'
    sheet['J1'] = 'Disponible'

    row = 2
    for teacher in teachers:
        sheet.cell(row=row, column=1).value = teacher.pk
        sheet.cell(row=row, column=2).value = teacher.user.username
        sheet.cell(row=row, column=3).value = teacher.first_name
        sheet.cell(row=row, column=4).value = teacher.last_name
        sheet.cell(row=row, column=5).value = teacher.gender
        date_formatted = teacher.birthday.strftime("%d-%m-%Y")
        sheet.cell(row=row, column=6).value = date_formatted
        sheet.cell(row=row, column=7).value = f"{teacher.address.country}-{teacher.address.city}-{teacher.address.street}"
        sheet.cell(row=row, column=8).value = teacher.phone_number
        sheet.cell(row=row, column=9).value = teacher.speciality 
        sheet.cell(row=row, column=10).value = 'OUI' if teacher.available else 'NON'
        
        row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Teacher_list_{datetime.today().strftime('%d_%m_%Y')}.xlsx"'

    workbook.save(response)
  
    return response


@login_required
def users_list_to_excel(request):
    users = UserModel.objects.all().order_by('id')

    workbook = Workbook()
    sheet = workbook.active

    sheet['A1'] = 'ID'
    sheet['B1'] = 'Nom d\'utilisateur'
    sheet['C1'] = 'Rôle'
    sheet['D1'] = 'Ecole'
    sheet['E1'] = 'Date de création'

    row = 2
    for user in users:
        roles = ', '.join([role.role for role in user.role.all()])
        sheet.cell(row=row, column=1).value = user.pk
        sheet.cell(row=row, column=2).value = user.username
        sheet.cell(row=row, column=3).value = roles
        sheet.cell(row=row, column=4).value = user.school.name
        sheet.cell(row=row, column=5).value = user.date_joined.strftime("%d-%m-%Y")
        
        row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="User_list_{datetime.today().strftime('%d_%m_%Y')}.xlsx"'

    workbook.save(response)
  
    return response


def generate_pdf(response, header, data):
    pdf = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    table = Table([header] + data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    
    table.setStyle(style)
    elements.append(table)

    pdf.build(elements)
    return response


@login_required
def students_list_to_pdf(request):
    students = StudentModel.objects.all().order_by('id')
    header = ['ID', 'Nom d\'utilisateur', 'Nom', 'Prénoms', 'Genre', 'Date de naissance', 'Adresse', 'Numéro de téléphone', 'Matricule', 'Numéro de téléphone du père']
    
    data = [
        [
            student.pk,
            student.user,
            student.first_name,
            student.last_name,
            student.gender,
            student.birthday.strftime("%d-%m-%Y"),
            student.address,
            student.phone_number,
            student.matricule,
            student.phone_number_father
        ]
        for student in students
    ]
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Student_list_{datetime.today().strftime('%d_%m_%Y')}.pdf"'
    
    return generate_pdf(response, header, data)


@login_required
def teachers_list_to_pdf(request):
    teachers = TeacherModel.objects.all().order_by('id')
    header = ['ID', 'Nom d\'utilisateur', 'Nom', 'Prénoms', 'Genre', 'Date de naissance', 'Adresse', 'Numéro de téléphone', 'Spécialité', 'Disponible']
    
    data = [
        [
            teacher.pk,
            teacher.user,
            teacher.first_name,
            teacher.last_name,
            teacher.gender,
            teacher.birthday.strftime("%d-%m-%Y"),
            teacher.address,
            teacher.phone_number,
            teacher.speciality,
            'OUI' if teacher.available else 'NON'
        ]
        for teacher in teachers
    ]
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Teacher_list_{datetime.today().strftime('%d_%m_%Y')}.pdf"'
    
    return generate_pdf(response, header, data)


@login_required
def users_list_to_pdf(request):
    users = UserModel.objects.all().order_by('id')
    header = ['ID', 'Nom d\'utilisateur', 'Rôle', 'Ecole', 'Date de création']

    data = []
    for user in users:
        roles = ', '.join([role.role for role in user.role.all()])
        data.append([
            user.pk,
            user.username,
            roles,
            user.school.name,
            user.date_joined.strftime("%d-%m-%Y")
        ])    
    

    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="User_list_{datetime.today().strftime('%d_%m_%Y')}.pdf"'
    
    return generate_pdf(response, header, data)